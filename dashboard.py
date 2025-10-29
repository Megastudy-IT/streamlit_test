import streamlit as st
from openpyxl import load_workbook, Workbook
from datetime import datetime
import matplotlib.pyplot as plt
import ianpack


# 엑셀 파일 경로
FILE_PATH = "inventory.xlsx"

# --------------------------
# 엑셀 데이터 불러오기 함수
# --------------------------
def load_data():
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    data = list(ws.values)[1:]
    return data, wb, ws

# --------------------------
# 엑셀에 신규 데이터 추가
# --------------------------
def add_item(product, category, stock, price, date):
    wb = load_workbook(FILE_PATH)
    ws = wb.active
    ws.append([product, category, stock, price, date])
    wb.save(FILE_PATH)

# --------------------------
# Streamlit 설정
# --------------------------
st.set_page_config(page_title="상품 재고 현황 대시보드", layout="wide")
st.title("상품 재고 현황 대시보드")
pack_sub = ianpack.welcome()
st.subheader(pack_sub)

# 탭 구성
tab1, tab2 = st.tabs(["재고 현황", " 신규 상품 등록"])

# -----------------------------------------------------------------
# 탭 1 : 재고 현황
# -----------------------------------------------------------------
with tab1:
    data, wb, ws = load_data()
    # 카테고리 목록 생성
# --------------------------
# 카테고리 목록 생성
# --------------------------
    category_set = set()
    for row in data:
        category_set.add(row[1])

    category_list = list(category_set)
    category_list.sort()
    categories = ["전체"] + category_list

    selected_category = st.sidebar.selectbox("카테고리 선택", categories)

    # --------------------------
    # 필터링
    # --------------------------
    filtered_data = []
    for row in data:
        if selected_category == "전체" or row[1] == selected_category: # row[1]의 예: 전자기기
            filtered_data.append(row)

    # --------------------------
    # 재고 5개 이하 상품 필터링
    # --------------------------
    low_stock = []
    for row in filtered_data:
        if row[2] <= 5:
            low_stock.append(row)

    # --------------------------
    # 총 재고 금액 계산
    # --------------------------
    total_value = 0
    for row in filtered_data:
        total_value += row[2] * row[3]


    # --------------------------
    # 요약 지표
    # --------------------------
    st.markdown("### 요약 정보")
    c1, c2, c3 = st.columns(3)
    c1.metric("총 상품 수", f"{len(filtered_data)} 개") # 폰트크기 고정
    c2.metric("총 재고 금액", f"{total_value:,} 원")
    c3.metric("재고 5개 이하 상품 수", f"{len(low_stock)} 개")

    # --------------------------
    # 재고 부족 상품 표시
    # --------------------------
    st.markdown("### 재고 5개 이하 상품")
    st.markdown("**상품명 | 카테고리 | 재고수량 | 단가 | 입고일**")
    if low_stock:
        st.table(low_stock)
    else:
        st.info("재고 부족 상품이 없습니다.")

    # --------------------------
    # 전체 목록 표시
    # --------------------------
    st.markdown("### 전체 상품 목록")
    st.markdown("**상품명 | 카테고리 | 재고수량 | 단가 | 입고일**")
    st.table(filtered_data)

    # --------------------------
    # 카테고리별 재고 시각화
    # --------------------------
    st.markdown("### 카테고리별 재고 현황")

    # 카테고리별 총 재고 계산
    category_summary = {}
    for row in data:
        cat = row[1]
        category_summary[cat] = category_summary.get(cat, 0) + row[2]

    fig, ax = plt.subplots(figsize=(4, 3), dpi=150)
    ax.bar(category_summary.keys(), category_summary.values(), color="skyblue", edgecolor="black")
    ax.set_ylabel("재고수량")
    ax.set_xlabel("카테고리")
    ax.set_title("카테고리별 총 재고수량")

    st.pyplot(fig, use_container_width=False)

# -----------------------------------------------------------------
# 탭 2 : 신규 상품 등록
# -----------------------------------------------------------------
with tab2:
    st.markdown("### 🆕 새 상품 입력")

    with st.form("add_form"):
        c1, c2 = st.columns(2)
        product = c1.text_input("상품명")
        category = c2.text_input("카테고리")

        c3, c4 = st.columns(2)
        stock = c3.number_input("재고수량", min_value=0, step=1)
        price = c4.number_input("단가(원)", min_value=0, step=100)

        date = st.date_input("입고일", datetime.today())
        submitted = st.form_submit_button("저장")

        if submitted:
            if product and category:
                add_item(product, category, stock, price, date.strftime("%Y-%m-%d"))
                st.success(f"{product} 상품이 추가되었습니다.")
            else:
                st.warning("상품명과 카테고리는 반드시 입력해야 합니다.")

# | 함수명                               | 주요 기능 설명                                                     |
# | --------------------------------- | ------------------------------------------------------------ |
# | **`st.set_page_config()`**        | 앱의 기본 설정을 지정합니다. 페이지 제목, 아이콘, 레이아웃(넓은 화면 등)을 설정할 수 있습니다.     |
# | **`st.title()`**                  | 페이지 상단에 큰 제목(Title)을 표시합니다.                                  |
# | **`st.tabs()`**                   | 여러 개의 탭(Tab) 인터페이스를 생성합니다. 각각의 탭 안에서 다른 UI나 데이터를 표시할 수 있습니다. |
# | **`st.sidebar.selectbox()`**      | 사이드바에 드롭다운 메뉴(선택 상자)를 표시합니다. 사용자가 선택한 값을 반환합니다.              |
# | **`st.markdown()`**               | 텍스트를 Markdown 형식으로 표시합니다. 제목, 강조, 리스트 등 다양한 서식 가능.           |
# | **`st.columns()`**                | 여러 개의 열(column)로 화면을 분할하여, 가로로 여러 요소를 배치할 수 있습니다.            |
# | **`st.metric()`**                 | 단일 지표(예: 합계, 평균, 변화율 등)를 시각적으로 보여줍니다.                        |
# | **`st.table()`**                  | 데이터를 표 형태로 렌더링합니다. 리스트, 딕셔너리, DataFrame 등을 입력할 수 있습니다.       |
# | **`st.info()`**                   | 정보성 메시지를 파란색 박스로 표시합니다.                                      |
# | **`st.pyplot()`**                 | Matplotlib으로 만든 그래프를 Streamlit 화면에 출력합니다.                    |
# | **`st.form()`**                   | 여러 입력 요소를 묶어 한 번에 제출할 수 있는 폼(Form) 영역을 생성합니다.                |
# | **`st.columns()`** *(폼 내부에서 재사용)* | 폼 내부에서도 입력 필드를 가로로 배치할 때 사용됩니다.                              |
# | **`st.text_input()`**             | 한 줄짜리 텍스트 입력창을 표시합니다.                                        |
# | **`st.number_input()`**           | 숫자 입력 위젯을 표시합니다. 최소값, 최대값, step 등을 설정 가능.                    |
# | **`st.date_input()`**             | 날짜를 선택할 수 있는 캘린더 입력 위젯을 표시합니다.                               |
# | **`st.form_submit_button()`**     | 폼 제출용 버튼을 생성합니다. 클릭 시 `True`를 반환하여 폼 제출을 감지합니다.              |
# | **`st.success()`**                | 녹색 박스로 성공 메시지를 표시합니다.                                        |
# | **`st.warning()`**                | 노란색 박스로 경고 메시지를 표시합니다.                                       |

